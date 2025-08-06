import requests
from bs4 import BeautifulSoup
import pandas as pd
from datetime import datetime
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ====== URLs ======
OUR_URL = "https://grillmaster.dp.ua/hazovi-hryli/"
COMPETITOR_URL = "https://bbq24.com.ua/ua/gazovye-grili/"

# ====== Заголовки для обхода блокировок ======
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/115.0 Safari/537.36"
    )
}

# ====== Парсер нашего сайта ======
def parse_grillmaster():
    response = requests.get(OUR_URL, headers=HEADERS)
    response.raise_for_status()
    soup = BeautifulSoup(response.text, "html.parser")

    products = {}
    for item in soup.select(".product"):
        title_tag = item.select_one(".woocommerce-loop-product__title")
        price_tag = item.select_one(".woocommerce-Price-amount.amount")
        if title_tag and price_tag:
            title = clean_name(title_tag.get_text(strip=True))
            price = parse_price(price_tag.get_text(strip=True))
            products[title] = price
    return products

# ====== Парсер BBQ24 ======
def parse_bbq24():
    response = requests.get(COMPETITOR_URL, headers=HEADERS)
    response.raise_for_status()
    soup = BeautifulSoup(response.text, "html.parser")

    products = {}
    for item in soup.select(".ty-grid-list__item"):
        title_tag = item.select_one(".ty-grid-list__item-name")
        price_tag = item.select_one(".ty-price")
        if title_tag and price_tag:
            title = clean_name(title_tag.get_text(strip=True))
            price = parse_price(price_tag.get_text(strip=True))
            products[title] = price
    return products

# ====== Очистка названия ======
def clean_name(name):
    name = name.lower()
    name = re.sub(r"[^a-zа-я0-9\s]", "", name)
    return name.strip()

# ====== Преобразование цены ======
def parse_price(price_text):
    digits = re.sub(r"[^\d]", "", price_text)
    return int(digits) if digits else None

# ====== Сравнение ======
def compare_prices(our_prices, competitor_prices):
    rows = []
    for our_name, our_price in our_prices.items():
        matched_competitor = None
        competitor_price = None

        for comp_name, comp_price in competitor_prices.items():
            # Проверяем вхождение хотя бы первых 2 слов
            if all(word in comp_name for word in our_name.split()[:2]):  
                matched_competitor = comp_name
                competitor_price = comp_price
                break

        if competitor_price is not None:
            diff = our_price - competitor_price
            rows.append([our_name, our_price, competitor_price, diff])
        else:
            rows.append([our_name, our_price, None, None])

    return rows

# ====== Сохранение в Excel с цветами ======
def save_to_excel(data):
    df = pd.DataFrame(data, columns=["Товар", "Наша цена", "Цена конкурента", "Разница"])
    filename = f"comparison_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    df.to_excel(filename, index=False)

    # Цветовая подсветка
    wb = load_workbook(filename)
    ws = wb.active

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # зелёный
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")    # красный
    gray_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")   # серый

    for row in range(2, ws.max_row + 1):
        diff = ws.cell(row=row, column=4).value
        if diff is not None:
            if diff < 0:  # мы дешевле
                ws.cell(row=row, column=4).fill = green_fill
            elif diff > 0:  # мы дороже
                ws.cell(row=row, column=4).fill = red_fill
            else:  # одинаково
                ws.cell(row=row, column=4).fill = gray_fill

    wb.save(filename)
    print(f"✅ Результат сохранён в {filename}")

# ====== Запуск ======
if __name__ == "__main__":
    print("📦 Парсим наш сайт...")
    our_prices = parse_grillmaster()
    print(f"  Найдено {len(our_prices)} товаров.")

    print("📦 Парсим сайт конкурента...")
    competitor_prices = parse_bbq24()
    print(f"  Найдено {len(competitor_prices)} товаров.")

    print("📊 Сравниваем цены...")
    comparison_data = compare_prices(our_prices, competitor_prices)

    print("💾 Сохраняем в Excel...")
    save_to_excel(comparison_data)
